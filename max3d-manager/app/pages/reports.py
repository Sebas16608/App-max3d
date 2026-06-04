import os
from PySide6.QtWidgets import (
    QWidget, QVBoxLayout, QHBoxLayout, QLabel, QPushButton,
    QMessageBox, QGroupBox, QFileDialog, QGridLayout, QFrame
)
from PySide6.QtCore import Qt
from app.utils import formatear_moneda


class ReportsPage(QWidget):
    def __init__(self, db, config, parent=None):
        super().__init__(parent)
        self.db = db
        self.config = config
        self.setup_ui()

    def setup_ui(self):
        layout = QVBoxLayout(self)
        layout.setContentsMargins(24, 24, 24, 24)
        layout.setSpacing(20)

        title = QLabel("Reportes")
        title.setStyleSheet("font-size: 28px; font-weight: bold; color: #e94560;")
        layout.addWidget(title)

        subtitle = QLabel("Genera reportes en PDF y Excel")
        subtitle.setStyleSheet("font-size: 14px; color: #a0a0b0; margin-bottom: 8px;")
        layout.addWidget(subtitle)

        pdf_group = QGroupBox("Reportes PDF")
        pdf_grid = QGridLayout(pdf_group)
        pdf_grid.setSpacing(12)

        pdf_reports = [
            ("Ventas", self.gen_pdf_ventas),
            ("Pedidos", self.gen_pdf_pedidos),
            ("Ganancias", self.gen_pdf_ganancias),
            ("Gastos", self.gen_pdf_gastos),
            ("Inventario", self.gen_pdf_inventario),
        ]

        for i, (name, func) in enumerate(pdf_reports):
            btn = QPushButton(f"📄 {name}")
            btn.setMinimumHeight(60)
            btn.setStyleSheet("font-size: 14px; font-weight: bold;")
            btn.clicked.connect(func)
            pdf_grid.addWidget(btn, i // 3, i % 3)

        layout.addWidget(pdf_group)

        excel_group = QGroupBox("Reportes Excel")
        excel_grid = QGridLayout(excel_group)
        excel_grid.setSpacing(12)

        excel_reports = [
            ("Ventas", self.gen_xlsx_ventas),
            ("Pedidos", self.gen_xlsx_pedidos),
            ("Ganancias", self.gen_xlsx_ganancias),
            ("Inventario", self.gen_xlsx_inventario),
            ("Inversiones", self.gen_xlsx_inversiones),
        ]

        for i, (name, func) in enumerate(excel_reports):
            btn = QPushButton(f"📊 {name}")
            btn.setMinimumHeight(60)
            btn.setStyleSheet("font-size: 14px; font-weight: bold;")
            btn.clicked.connect(func)
            excel_grid.addWidget(btn, i // 3, i % 3)

        layout.addWidget(excel_group)

        layout.addStretch()

    def get_save_path(self, default_name, ext):
        return QFileDialog.getSaveFileName(self, "Guardar reporte", os.path.expanduser(f"~/{default_name}"), f"*.{ext}")[0]

    def gen_pdf_ventas(self):
        path = self.get_save_path("reporte_ventas.pdf", "pdf")
        if not path:
            return
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas
            from reportlab.lib import colors

            c = canvas.Canvas(path, pagesize=letter)
            w, h = letter
            c.setTitle("Reporte de Ventas - Max3D Manager")

            c.setFont("Helvetica-Bold", 20)
            c.drawString(40, h - 40, "Reporte de Ventas")

            c.setFont("Helvetica", 10)
            c.drawString(40, h - 60, f"Generado: {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M')}")

            ventas = self.db.fetch_all("""
                SELECT cp.*, c.nombre as cliente_nombre,
                    COALESCE(p.nombre, cp.nombre_personalizado) as prod_nombre
                FROM clientes_productos cp
                JOIN clientes c ON cp.cliente_id = c.id
                LEFT JOIN productos p ON cp.producto_id = p.id
                WHERE cp.estado = 'Entregado'
                ORDER BY cp.fecha_pedido DESC
            """)

            y = h - 90
            c.setFont("Helvetica-Bold", 11)
            c.drawString(40, y, "Cliente")
            c.drawString(200, y, "Producto")
            c.drawString(360, y, "Total")
            c.drawString(440, y, "Ganancia")
            y -= 20

            c.setFont("Helvetica", 10)
            total_ventas = 0
            total_ganancia = 0
            for v in ventas:
                if y < 50:
                    c.showPage()
                    y = h - 50
                c.drawString(40, y, v["cliente_nombre"][:25])
                c.drawString(200, y, (v["prod_nombre"] or "—")[:25])
                c.drawString(360, y, f"Q{v['precio_final']:.2f}")
                gan = (v["precio_final"] or 0) - (v["costo_total"] or 0)
                c.drawString(440, y, f"Q{gan:.2f}")
                total_ventas += v["precio_final"] or 0
                total_ganancia += gan
                y -= 16

            y -= 10
            c.setFont("Helvetica-Bold", 12)
            c.drawString(40, y, f"Total Ventas: Q{total_ventas:.2f}")
            c.drawString(300, y, f"Ganancia Total: Q{total_ganancia:.2f}")

            c.save()
            QMessageBox.information(self, "Éxito", f"PDF generado: {path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al generar PDF: {str(e)}")

    def gen_pdf_pedidos(self):
        path = self.get_save_path("reporte_pedidos.pdf", "pdf")
        if not path:
            return
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas

            c = canvas.Canvas(path, pagesize=letter)
            w, h = letter
            c.setTitle("Reporte de Pedidos - Max3D Manager")

            c.setFont("Helvetica-Bold", 20)
            c.drawString(40, h - 40, "Reporte de Pedidos")

            c.setFont("Helvetica", 10)
            c.drawString(40, h - 60, f"Generado: {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M')}")

            pedidos = self.db.fetch_all("""
                SELECT cp.*, c.nombre as cliente_nombre,
                    COALESCE(p.nombre, cp.nombre_personalizado) as prod_nombre
                FROM clientes_productos cp
                JOIN clientes c ON cp.cliente_id = c.id
                LEFT JOIN productos p ON cp.producto_id = p.id
                ORDER BY cp.fecha_pedido DESC
            """)

            y = h - 90
            c.setFont("Helvetica-Bold", 9)
            c.drawString(40, y, "Cliente")
            c.drawString(180, y, "Producto")
            c.drawString(320, y, "Estado")
            c.drawString(390, y, "Total")
            c.drawString(460, y, "Fecha")
            y -= 18

            c.setFont("Helvetica", 9)
            for p in pedidos:
                if y < 50:
                    c.showPage()
                    y = h - 50
                c.drawString(40, y, p["cliente_nombre"][:22])
                c.drawString(180, y, (p["prod_nombre"] or "—")[:22])
                c.drawString(320, y, p["estado"])
                c.drawString(390, y, f"Q{p['precio_final']:.2f}" if p["precio_final"] else "—")
                c.drawString(460, y, p["fecha_pedido"][:10] if p["fecha_pedido"] else "")
                y -= 14

            c.save()
            QMessageBox.information(self, "Éxito", f"PDF generado: {path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al generar PDF: {str(e)}")

    def gen_pdf_ganancias(self):
        path = self.get_save_path("reporte_ganancias.pdf", "pdf")
        if not path:
            return
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas

            c = canvas.Canvas(path, pagesize=letter)
            w, h = letter
            c.setTitle("Reporte de Ganancias - Max3D Manager")

            c.setFont("Helvetica-Bold", 20)
            c.drawString(40, h - 40, "Reporte de Ganancias")

            c.setFont("Helvetica", 10)
            c.drawString(40, h - 60, f"Generado: {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M')}")

            ventas = self.db.fetch_one("SELECT COALESCE(SUM(precio_final),0) as total FROM clientes_productos WHERE estado='Entregado'")
            total_ventas = ventas["total"] if ventas else 0

            costos = self.db.fetch_one("SELECT COALESCE(SUM(costo_total),0) as total FROM clientes_productos WHERE estado='Entregado'")
            total_costos = costos["total"] if costos else 0

            gastos = self.db.fetch_one("SELECT COALESCE(SUM(monto),0) as total FROM inversiones")
            total_gastos = gastos["total"] if gastos else 0

            ganancia_bruta = total_ventas - total_costos
            ganancia_neta = ganancia_bruta - total_gastos

            y = h - 100
            c.setFont("Helvetica", 14)
            c.drawString(40, y, f"Ventas totales: Q{total_ventas:.2f}")
            y -= 25
            c.drawString(40, y, f"Costos totales: Q{total_costos:.2f}")
            y -= 25
            c.drawString(40, y, f"Ganancia bruta: Q{ganancia_bruta:.2f}")
            y -= 25
            c.drawString(40, y, f"Gastos totales: Q{total_gastos:.2f}")
            y -= 25
            c.setFont("Helvetica-Bold", 16)
            c.drawString(40, y, f"Ganancia neta: Q{ganancia_neta:.2f}")

            c.save()
            QMessageBox.information(self, "Éxito", f"PDF generado: {path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al generar PDF: {str(e)}")

    def gen_pdf_gastos(self):
        path = self.get_save_path("reporte_gastos.pdf", "pdf")
        if not path:
            return
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas

            c = canvas.Canvas(path, pagesize=letter)
            w, h = letter
            c.setTitle("Reporte de Gastos - Max3D Manager")

            c.setFont("Helvetica-Bold", 20)
            c.drawString(40, h - 40, "Reporte de Gastos")

            c.setFont("Helvetica", 10)
            c.drawString(40, h - 60, f"Generado: {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M')}")

            gastos = self.db.fetch_all("SELECT * FROM inversiones ORDER BY fecha DESC")

            y = h - 90
            c.setFont("Helvetica-Bold", 11)
            c.drawString(40, y, "Fecha")
            c.drawString(140, y, "Categoría")
            c.drawString(280, y, "Descripción")
            c.drawString(440, y, "Monto")
            y -= 20

            c.setFont("Helvetica", 10)
            total = 0
            for g in gastos:
                if y < 50:
                    c.showPage()
                    y = h - 50
                c.drawString(40, y, g["fecha"] or "")
                c.drawString(140, y, g["categoria"])
                c.drawString(280, y, (g["descripcion"] or "")[:25])
                c.drawString(440, y, f"Q{g['monto']:.2f}")
                total += g["monto"]
                y -= 16

            y -= 10
            c.setFont("Helvetica-Bold", 12)
            c.drawString(40, y, f"Total gastos: Q{total:.2f}")

            c.save()
            QMessageBox.information(self, "Éxito", f"PDF generado: {path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al generar PDF: {str(e)}")

    def gen_pdf_inventario(self):
        path = self.get_save_path("reporte_inventario.pdf", "pdf")
        if not path:
            return
        try:
            from reportlab.lib.pagesizes import letter
            from reportlab.pdfgen import canvas

            c = canvas.Canvas(path, pagesize=letter)
            w, h = letter
            c.setTitle("Reporte de Inventario - Max3D Manager")

            c.setFont("Helvetica-Bold", 20)
            c.drawString(40, h - 40, "Reporte de Inventario")

            c.setFont("Helvetica", 10)
            c.drawString(40, h - 60, f"Generado: {__import__('datetime').datetime.now().strftime('%Y-%m-%d %H:%M')}")

            y = h - 90
            c.setFont("Helvetica-Bold", 14)
            c.drawString(40, y, "Filamentos")
            y -= 20

            filamentos = self.db.fetch_all("SELECT * FROM filamentos ORDER BY nombre")
            c.setFont("Helvetica", 10)
            c.drawString(40, y, "Nombre")
            c.drawString(200, y, "Tipo")
            c.drawString(300, y, "Color")
            c.drawString(400, y, "Peso disponible")
            y -= 16

            for f in filamentos:
                if y < 50:
                    c.showPage()
                    y = h - 50
                c.drawString(40, y, f["nombre"][:25])
                c.drawString(200, y, f["tipo"] or "")
                c.drawString(300, y, f["color"] or "")
                c.drawString(400, y, f"{f['peso_disponible']:.0f}g")
                y -= 14

            y -= 10
            c.setFont("Helvetica-Bold", 14)
            c.drawString(40, y, "Pinturas")
            y -= 20

            pinturas = self.db.fetch_all("SELECT * FROM pinturas ORDER BY color")
            c.setFont("Helvetica", 10)
            c.drawString(40, y, "Color")
            c.drawString(200, y, "Marca")
            c.drawString(350, y, "Cantidad")
            y -= 16

            for p in pinturas:
                if y < 50:
                    c.showPage()
                    y = h - 50
                c.drawString(40, y, p["color"] or "")
                c.drawString(200, y, p["marca"] or "")
                c.drawString(350, y, str(p["cantidad_disponible"]))
                y -= 14

            c.save()
            QMessageBox.information(self, "Éxito", f"PDF generado: {path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al generar PDF: {str(e)}")

    def gen_xlsx_ventas(self):
        path = self.get_save_path("reporte_ventas.xlsx", "xlsx")
        if not path:
            return
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Ventas"

            ws.append(["Cliente", "Producto", "Cantidad", "Estado", "Total", "Costo", "Ganancia", "Fecha"])

            ventas = self.db.fetch_all("""
                SELECT cp.*, c.nombre as cliente_nombre,
                    COALESCE(p.nombre, cp.nombre_personalizado) as prod_nombre
                FROM clientes_productos cp
                JOIN clientes c ON cp.cliente_id = c.id
                LEFT JOIN productos p ON cp.producto_id = p.id
                WHERE cp.estado = 'Entregado'
                ORDER BY cp.fecha_pedido DESC
            """)

            for v in ventas:
                gan = (v["precio_final"] or 0) - (v["costo_total"] or 0)
                ws.append([
                    v["cliente_nombre"],
                    v["prod_nombre"],
                    v["cantidad"],
                    v["estado"],
                    v["precio_final"] or 0,
                    v["costo_total"] or 0,
                    gan,
                    v["fecha_pedido"][:10] if v["fecha_pedido"] else ""
                ])

            wb.save(path)
            QMessageBox.information(self, "Éxito", f"Excel generado: {path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al generar Excel: {str(e)}")

    def gen_xlsx_pedidos(self):
        path = self.get_save_path("reporte_pedidos.xlsx", "xlsx")
        if not path:
            return
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Pedidos"

            ws.append(["Cliente", "Producto", "Cantidad", "Estado", "Total", "Pagado", "Pendiente", "Fecha"])

            pedidos = self.db.fetch_all("""
                SELECT cp.*, c.nombre as cliente_nombre,
                    COALESCE(p.nombre, cp.nombre_personalizado) as prod_nombre
                FROM clientes_productos cp
                JOIN clientes c ON cp.cliente_id = c.id
                LEFT JOIN productos p ON cp.producto_id = p.id
                ORDER BY cp.fecha_pedido DESC
            """)

            for p in pedidos:
                pendiente = (p["precio_final"] or 0) - (p["pago_total"] or 0)
                ws.append([
                    p["cliente_nombre"],
                    p["prod_nombre"],
                    p["cantidad"],
                    p["estado"],
                    p["precio_final"] or 0,
                    p["pago_total"] or 0,
                    pendiente,
                    p["fecha_pedido"][:10] if p["fecha_pedido"] else ""
                ])

            wb.save(path)
            QMessageBox.information(self, "Éxito", f"Excel generado: {path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al generar Excel: {str(e)}")

    def gen_xlsx_ganancias(self):
        path = self.get_save_path("reporte_ganancias.xlsx", "xlsx")
        if not path:
            return
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Ganancias"

            ws.append(["Producto", "Cliente", "Total Venta", "Costo Total", "Ganancia", "Fecha"])

            data = self.db.fetch_all("""
                SELECT cp.*, c.nombre as cliente_nombre,
                    COALESCE(p.nombre, cp.nombre_personalizado) as prod_nombre
                FROM clientes_productos cp
                JOIN clientes c ON cp.cliente_id = c.id
                LEFT JOIN productos p ON cp.producto_id = p.id
                WHERE cp.estado = 'Entregado'
                ORDER BY cp.fecha_pedido DESC
            """)

            for d in data:
                gan = (d["precio_final"] or 0) - (d["costo_total"] or 0)
                ws.append([
                    d["prod_nombre"],
                    d["cliente_nombre"],
                    d["precio_final"] or 0,
                    d["costo_total"] or 0,
                    gan,
                    d["fecha_pedido"][:10] if d["fecha_pedido"] else ""
                ])

            wb.save(path)
            QMessageBox.information(self, "Éxito", f"Excel generado: {path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al generar Excel: {str(e)}")

    def gen_xlsx_inventario(self):
        path = self.get_save_path("reporte_inventario.xlsx", "xlsx")
        if not path:
            return
        try:
            from openpyxl import Workbook
            wb = Workbook()

            ws_fil = wb.active
            ws_fil.title = "Filamentos"
            ws_fil.append(["Nombre", "Marca", "Tipo", "Color", "Peso disponible (g)", "Precio compra", "Precio/g"])
            for f in self.db.fetch_all("SELECT * FROM filamentos ORDER BY nombre"):
                ws_fil.append([f["nombre"], f["marca"], f["tipo"], f["color"], f["peso_disponible"], f["precio_compra"], f["precio_por_gramo"]])

            ws_pin = wb.create_sheet("Pinturas")
            ws_pin.append(["Marca", "Tipo", "Color", "Cantidad", "Precio compra"])
            for p in self.db.fetch_all("SELECT * FROM pinturas ORDER BY color"):
                ws_pin.append([p["marca"], p["tipo"], p["color"], p["cantidad_disponible"], p["precio_compra"]])

            ws_otro = wb.create_sheet("Otros Materiales")
            ws_otro.append(["Nombre", "Cantidad", "Costo"])
            for o in self.db.fetch_all("SELECT * FROM otros_materiales ORDER BY nombre"):
                ws_otro.append([o["nombre"], o["cantidad"], o["costo"]])

            wb.save(path)
            QMessageBox.information(self, "Éxito", f"Excel generado: {path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al generar Excel: {str(e)}")

    def gen_xlsx_inversiones(self):
        path = self.get_save_path("reporte_inversiones.xlsx", "xlsx")
        if not path:
            return
        try:
            from openpyxl import Workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Inversiones"

            ws.append(["Fecha", "Categoría", "Descripción", "Monto"])
            for inv in self.db.fetch_all("SELECT * FROM inversiones ORDER BY fecha DESC"):
                ws.append([inv["fecha"], inv["categoria"], inv["descripcion"], inv["monto"]])

            wb.save(path)
            QMessageBox.information(self, "Éxito", f"Excel generado: {path}")
        except Exception as e:
            QMessageBox.critical(self, "Error", f"Error al generar Excel: {str(e)}")
